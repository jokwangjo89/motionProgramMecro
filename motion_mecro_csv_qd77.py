import chardet
from openpyxl import Workbook,load_workbook
from mecro import MotionConfigurator,add_to_address,add_to_bit_address,add_bit_offset,add_hex_offset
excel_file ='QD77_MOTION_MECRO_MAP_SAMPLE.xlsx'
motion = MotionConfigurator(excel_file)




################################# 2. D6부터 아래로 채울 데이터################################################
#wb = Workbook()
wb= load_workbook("motion_mecro_qd77_result2.xlsx")
ws = wb.active
ws.title = "Exported"
data = []
last_row = 2
# B열 = 2번째 열, D열 = 4번째 열
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=5):
    for cell in row:
        if cell.column == 3 or cell.column == 5 or cell.column == 1:
        #if cell.column == 3 or cell.column == 5 :
            cell.value = None
data.append("")
for j in range(10):
    
    # 1. 워크북 및 워크시트 생성
    
    program_no= 18
    program_count= j+1
    axis_no = motion.motor_prog.get_axis_no(program_no, program_count)
    scan_cnt = j
    axis_address_offset = 20 *scan_cnt
    
    if str(axis_no).isdigit():
        qd77_axis_start = motion.motor_param_adr_qd77.get_start_axis('START_AXIS',axis_no)-1
        cell = f'A{last_row}'
        ws[cell] = 'State'
        last_row=last_row+1
        data.append("")
        #MOTOR STATE
        last_row=motion.ladder_mecro_qd77.export_partial_to_excel(ws,last_row,"STATE") 
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),1+scan_cnt))
        data.append("")
        data.append(add_bit_offset(motion.motor_param_adr_qd77.get_param_adr('EXT_IO_ADR',axis_no),1))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+axis_address_offset),0))
        data.append("")
        data.append(add_bit_offset(motion.motor_param_adr_qd77.get_param_adr('EXT_IO_ADR',axis_no),0))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+axis_address_offset),1))
        data.append("")
        data.append(add_bit_offset(motion.motor_param_adr_qd77.get_param_adr('EXT_IO_ADR',axis_no),6))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+axis_address_offset),2))
        data.append("")
        data.append(add_bit_offset(add_to_address(motion.motor_param_adr_qd77.get_param_adr('EXT_IO_ADR',axis_no),1),13))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+axis_address_offset),3))
        data.append("")
        data.append("SM401")
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+axis_address_offset),4))
        data.append("")
        data.append(add_bit_offset(add_to_address(motion.motor_param_adr_qd77.get_param_adr('EXT_IO_ADR',axis_no),1),3))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+axis_address_offset),5))
        data.append("")
        data.append(add_bit_offset(add_to_address(motion.motor_param_adr_qd77.get_param_adr('EXT_IO_ADR',axis_no),1),4))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+axis_address_offset),6))
        data.append("")
        data.append(add_bit_offset(add_to_address(motion.motor_param_adr_qd77.get_param_adr('STATUS_ADR',axis_no),0),1))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+axis_address_offset),7))
        data.append("")
        data.append("F13")
        data.append(add_hex_offset(motion.motor_param_adr_qd77.get_param_adr('START_X_ADR',axis_no,qd77_axis_start),16))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+axis_address_offset),8))
        data.append("")
        data.append("F13")
        data.append(add_hex_offset(motion.motor_param_adr_qd77.get_param_adr('START_X_ADR',axis_no,qd77_axis_start),16))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+axis_address_offset),9))
        
        cell = f'A{last_row}'
        ws[cell] = 'Origin'
        last_row=last_row+1
        #MOTOR ORIGIN
        last_row=motion.ladder_mecro_qd77.export_partial_to_excel(ws,last_row,"ORIGIN") 
        data.append("")
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),1+scan_cnt))
        data.append("F10")
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),20+axis_address_offset))
        data.append("")
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+axis_address_offset),8))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+axis_address_offset),9))
        data.append("F10")
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('T_START',program_no),20+axis_address_offset))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),20+axis_address_offset))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('T_START',program_no),20+axis_address_offset))
        data.append("K30")
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),1+scan_cnt))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),20+axis_address_offset))
        data.append("SM412")
        data.append("F11")
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),1+scan_cnt))
        data.append("F10")
        data.append("L12")
        data.append("")
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+8+axis_address_offset),0))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),1+scan_cnt))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),20+axis_address_offset))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+1+axis_address_offset),0))
        data.append("")
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+axis_address_offset),3))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+axis_address_offset),4))
        data.append(add_to_address('M600',axis_no-1))
        data.append("")
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),20+5+axis_address_offset))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+8+axis_address_offset),0))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),20+2+axis_address_offset))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),1+scan_cnt))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+axis_address_offset),5))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+axis_address_offset),6))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),20+3+axis_address_offset))
        data.append("")
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+axis_address_offset),7))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),20+3+axis_address_offset))
        cell = f'A{last_row}'
        ws[cell] = 'POSITION_INTERLOCK'
        last_row=last_row+1
        data.append("")
        #POSITION_INTERLOCK
        last_row=motion.ladder_mecro_qd77.export_partial_to_excel(ws,last_row,"POSITION_INTERLOCK")  
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),1+scan_cnt))
        data.append("")
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),20+5+axis_address_offset))
        data.append("")
        data.append("F10")
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+ 2 +axis_address_offset),0))
        data.append("")
        data.append("F10")
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+ 4 +axis_address_offset),0))
        data.append("")
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),20+3+axis_address_offset))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+ 2 +axis_address_offset),0))
        data.append("F10")
        for i in range(1, 16):
            data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+ 2 +axis_address_offset),i))
            i= i + 1
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),1+scan_cnt))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),20+3+axis_address_offset))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+ 4 +axis_address_offset),0))
        data.append("F10")
        for i in range(1, 16):
            data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+ 4 +axis_address_offset),i))
            i= i + 1
        cell = f'A{last_row}'
        ws[cell] = 'POSITION_BIT'
        last_row=last_row+1
        data.append("")
        #POSITION_BIT
        last_row=motion.ladder_mecro_qd77.export_partial_to_excel(ws,last_row,"POSITION_BIT")  
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),1+scan_cnt))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),20+2+axis_address_offset))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+ 4 +axis_address_offset),0))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+ 6 +axis_address_offset),0))
        for i in range(1, 16):
            data.append("F10")
            data.append("")
            data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+ 2 +axis_address_offset),i))
            data.append("F10")
            data.append("")
            data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+ 4 +axis_address_offset),i))
            data.append("")
            data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+ 6 +axis_address_offset),i))
            i= i + 1
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),1+scan_cnt))
        data.append("K0")
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+6+axis_address_offset))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),20+4+axis_address_offset))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),1+scan_cnt))
        data.append("")
        data.append(add_hex_offset(motion.motor_param_adr_qd77.get_param_adr('START_Y_ADR',axis_no,qd77_axis_start),4))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+ 0 +axis_address_offset),3))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+ 0 +axis_address_offset),4))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+ 0 +axis_address_offset),8))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+ 0 +axis_address_offset),9))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+ 0 +axis_address_offset),7))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),20+5+axis_address_offset))
        data.append("")
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),20+4+axis_address_offset))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),20+5+axis_address_offset))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),20+6+axis_address_offset))

        cell = f'A{last_row}'
        ws[cell] = 'POSITION_START'
        last_row=last_row+1
        data.append("")
        #POSITION_START
        last_row=motion.ladder_mecro_qd77.export_partial_to_excel(ws,last_row,"POSITION_START")  
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),1+scan_cnt))
        data.append("")
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+ 19 +axis_address_offset),0))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+ 19 +axis_address_offset),1))
        data.append("")
        data.append("K0")
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+10+axis_address_offset))
        data.append("")
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),20+4+axis_address_offset))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+ 6 +axis_address_offset),0))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+ 11 +axis_address_offset),0))
        data.append("K9001")
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+10+axis_address_offset))
        data.append("")
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+ 0 +axis_address_offset),3))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+ 0 +axis_address_offset),4))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+ 0 +axis_address_offset),8))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+ 0 +axis_address_offset),9))
        data.append("")
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+ 11 +axis_address_offset),0))
        for i in range(0, 16):
            data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),20+6+axis_address_offset))
            data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+ 6 +axis_address_offset),i))
            if i == 0:
                data.append("K9001")
            if i != 0:
                data.append(add_to_address('K0',i))
            data.append(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+10+axis_address_offset))
            i = i + 1
        
        cell = f'A{last_row}'
        ws[cell] = 'POSITION_MOVEMENT'
        last_row=last_row+1
        data.append("")
        #POSITION_MOVEMENT
        last_row=motion.ladder_mecro_qd77.export_partial_to_excel(ws,last_row,"POSITION_MOVEMENT")  
        data.append("K0")
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+6+axis_address_offset))
        data.append("K0")
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+7+axis_address_offset))
        data.append("K1")
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+10+axis_address_offset))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),20+10+axis_address_offset))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),1+scan_cnt))
        for i in range(1, 16):
            data.append("K0")
            data.append(motion.motor_pos_adr.get_pos_adr('MODEL',axis_no, i))
            i = i + 1
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),1+scan_cnt))
        for i in range(1, 16):
            data.append("K0")
            data.append(motion.motor_pos_adr.get_pos_adr('MACHINE',axis_no, i))
            i = i + 1
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),1+scan_cnt))
        for i in range(1, 16):
            data.append("K0")
            data.append(motion.motor_pos_adr.get_pos_adr('VISION',axis_no, i))
            i = i + 1
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),1+scan_cnt))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),20+10+axis_address_offset))
        data.append("")
        data.append("")
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+ 11 +axis_address_offset),0))
        data.append("K9001")
        data.append(f"{motion.motor_param_adr_qd77.get_param_adr('U_START',axis_no)}\{motion.motor_param_adr_qd77.get_param_adr('POSNO_B_ADR',axis_no,qd77_axis_start)}")
        data.append("")
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+11+axis_address_offset),0))
        data.append("K1")
        data.append(f"{motion.motor_param_adr_qd77.get_param_adr('U_START',axis_no)}\{motion.motor_param_adr_qd77.get_param_adr('POSNO_B_ADR',axis_no,qd77_axis_start)}")
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+10+axis_address_offset))
        data.append("K1")
        data.append("Z4")
        data.append("Z4")
        data.append("K2")
        data.append("Z4")
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+10+axis_address_offset))
        data.append(f"{motion.motor_param_adr_qd77.get_param_adr('U_START',axis_no)}\{motion.motor_param_adr_qd77.get_param_adr('MCODE_B_ADR',axis_no,qd77_axis_start)}")
        data.append("")
        data.append("SM400")
        data.append(motion.motor_param_adr.get_param_adr('ACCTIME',axis_no))
        data.append(f"{motion.motor_param_adr_qd77.get_param_adr('U_START',axis_no)}\{motion.motor_param_adr_qd77.get_param_adr('ACC_B_ADR',axis_no,qd77_axis_start)}")
        data.append(motion.motor_param_adr.get_param_adr('DECTIME',axis_no))
        data.append(f"{motion.motor_param_adr_qd77.get_param_adr('U_START',axis_no)}\{motion.motor_param_adr_qd77.get_param_adr('DEC_B_ADR',axis_no,qd77_axis_start)}")
        data.append("")
        data.append("SM400")
        data.append("")
        data.append("L10")
        data.append(motion.motor_param_adr.get_param_adr('HIGHSPEED',axis_no))
        data.append("ZR180")
        data.append("ZR186")
        data.append("")
        data.append("L10")
        data.append(motion.motor_pos_adr.get_pos_adr('INDSPEED',axis_no,1)+'Z4')
        data.append("ZR180")
        data.append("ZR186")
        data.append("")
        data.append("SM400")
        data.append("ZR186")
        data.append(f"{motion.motor_param_adr_qd77.get_param_adr('U_START',axis_no)}\{motion.motor_param_adr_qd77.get_param_adr('SPEED_B_ADR',axis_no,qd77_axis_start)}")
        data.append(f"{motion.motor_param_adr_qd77.get_param_adr('U_START',axis_no)}\{motion.motor_param_adr_qd77.get_param_adr('SPEED_B_ADR',axis_no,qd77_axis_start)}")
        data.append(f"{motion.motor_param_adr_qd77.get_param_adr('U_START',axis_no)}\{motion.motor_param_adr_qd77.get_param_adr('A_SPDLIMIT_B_ADR',axis_no,qd77_axis_start)}")
        data.append("")
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+11+axis_address_offset),0))
        data.append(motion.motor_pos_adr.get_pos_adr('FIXED',axis_no, 1)+'Z4')
        data.append(motion.motor_pos_adr.get_pos_adr('OFFSET',axis_no, 1)+'Z4')
        data.append(motion.motor_pos_adr.get_pos_adr('TARGET',axis_no, 1)+'Z4')
        data.append(motion.motor_pos_adr.get_pos_adr('TARGET',axis_no, 1)+'Z4')
        data.append(motion.motor_pos_adr.get_pos_adr('MODEL',axis_no, 1)+'Z4')
        data.append(motion.motor_pos_adr.get_pos_adr('TARGET',axis_no, 1)+'Z4')
        data.append(motion.motor_pos_adr.get_pos_adr('TARGET',axis_no, 1)+'Z4')
        data.append(motion.motor_pos_adr.get_pos_adr('MACHINE',axis_no, 1)+'Z4')
        data.append(motion.motor_pos_adr.get_pos_adr('TARGET',axis_no, 1)+'Z4')
        data.append(motion.motor_pos_adr.get_pos_adr('TARGET',axis_no, 1)+'Z4')
        data.append(motion.motor_pos_adr.get_pos_adr('VISION',axis_no, 1)+'Z4')
        data.append(motion.motor_pos_adr.get_pos_adr('TARGET',axis_no, 1)+'Z4')
        data.append(motion.motor_pos_adr.get_pos_adr('TARGET',axis_no, 1)+'Z4')
        data.append(f"{motion.motor_param_adr_qd77.get_param_adr('U_START',axis_no)}\{motion.motor_param_adr_qd77.get_param_adr('POSADR_B_ADR',axis_no,qd77_axis_start)}")
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),1+scan_cnt))
        data.append("")
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),20+10+axis_address_offset))
        data.append("")
        data.append("L11")
        data.append(motion.motor_param_adr.get_param_adr('SOFTLIMIT+',axis_no))
        data.append(f"{motion.motor_param_adr_qd77.get_param_adr('U_START',axis_no)}\{motion.motor_param_adr_qd77.get_param_adr('POSADR_B_ADR',axis_no,qd77_axis_start)}")
        data.append(motion.motor_param_adr.get_param_adr('SOFTLIMIT-',axis_no))
        data.append(f"{motion.motor_param_adr_qd77.get_param_adr('U_START',axis_no)}\{motion.motor_param_adr_qd77.get_param_adr('POSADR_B_ADR',axis_no,qd77_axis_start)}")
        data.append("")
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),20+11+axis_address_offset))
        data.append("")
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),20+11+axis_address_offset))
        data.append(add_hex_offset(motion.motor_param_adr_qd77.get_param_adr('START_Y_ADR',axis_no,qd77_axis_start),16))
        data.append("")
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+axis_address_offset),3))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+axis_address_offset),4))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+axis_address_offset),8))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+axis_address_offset),9))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),20+11+axis_address_offset))
        data.append("")
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),20+10+axis_address_offset))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),1+scan_cnt))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+axis_address_offset),3))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+axis_address_offset),4))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),20+7+axis_address_offset))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),1+scan_cnt))
        data.append("")
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+11+axis_address_offset),0))
        data.append(add_hex_offset(motion.motor_param_adr_qd77.get_param_adr('START_Y_ADR',axis_no,qd77_axis_start),16))
        data.append(add_hex_offset(motion.motor_param_adr_qd77.get_param_adr('JOG_Y_ADR',axis_no,qd77_axis_start),8))
        data.append(add_hex_offset(motion.motor_param_adr_qd77.get_param_adr('JOG_Y_ADR',axis_no,qd77_axis_start),9))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+axis_address_offset),3))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+axis_address_offset),4))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),20+3+axis_address_offset))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+axis_address_offset),8))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+axis_address_offset),9))
        data.append("")
        data.append("K0")
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+12+axis_address_offset))
        data.append("")
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+axis_address_offset),9))
        data.append("")
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),20+7+axis_address_offset))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+10+axis_address_offset))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+12+axis_address_offset))
        
        cell = f'A{last_row}'
        ws[cell] = 'POSITION_CHECK'
        last_row=last_row+1
        data.append("")
        #POSITION_CHECK
        last_row=motion.ladder_mecro_qd77.export_partial_to_excel(ws,last_row,"POSITION_CHECK")  
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),1+scan_cnt))
        data.append(f"{motion.motor_param_adr_qd77.get_param_adr('U_START',axis_no)}\{motion.motor_param_adr_qd77.get_param_adr('POSADR_B_ADR',axis_no,qd77_axis_start)}")
        data.append("K10")
        data.append("D800")
        data.append(f"{motion.motor_param_adr_qd77.get_param_adr('U_START',axis_no)}\{motion.motor_param_adr_qd77.get_param_adr('POSADR_B_ADR',axis_no,qd77_axis_start)}")
        data.append("K10")
        data.append("D802")
        data.append(motion.motor_param_adr.get_param_adr('CURPOSITION',axis_no))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+14+axis_address_offset))
        data.append(motion.motor_param_adr.get_param_adr('CURPOSITION',axis_no))
        data.append("D800")
        data.append(motion.motor_param_adr.get_param_adr('CURPOSITION',axis_no))
        data.append("D802")
        data.append(f"{motion.motor_param_adr_qd77.get_param_adr('U_START',axis_no)}\{motion.motor_param_adr_qd77.get_param_adr('POSADR_B_ADR',axis_no,qd77_axis_start)}")
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+14+axis_address_offset))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),1+scan_cnt))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),20+7+axis_address_offset))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),20+3+axis_address_offset))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+axis_address_offset),8))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+axis_address_offset),9))

        for i in range(1, 16):
            data.append("")
            data.append(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+12+axis_address_offset))
            data.append(add_to_address('K0',i))
            data.append(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+14+axis_address_offset))
            data.append(motion.motor_pos_adr.get_pos_adr('TARGET',axis_no, i))
            data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+16+axis_address_offset),i-1))
        
        cell = f'A{last_row}'
        ws[cell] = 'MANUAL_COMMAND'
        last_row=last_row+1
        data.append("")
        #MANUAL_CAMMAND
        last_row=motion.ladder_mecro_qd77.export_partial_to_excel(ws,last_row,"MANUAL_COMMAND")  
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),1+scan_cnt))
        data.append("")
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+18+axis_address_offset))
        data.append("K0")
        data.append(motion.motor_param_adr.get_param_adr('JOGLOWSPD',axis_no))
        data.append("ZR190")
        data.append("ZR192")
        data.append("")
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+18+axis_address_offset))
        data.append("K1")
        data.append(motion.motor_param_adr.get_param_adr('JOGMIDSPD',axis_no))
        data.append("ZR190")
        data.append("ZR192")
        data.append("")
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+18+axis_address_offset))
        data.append("K2")
        data.append(motion.motor_param_adr.get_param_adr('JOGHIGHSPD',axis_no))
        data.append("ZR190")
        data.append("ZR192")
        data.append("")
        data.append("ZR192")
        data.append(f"{motion.motor_param_adr_qd77.get_param_adr('U_START',axis_no)}\{motion.motor_param_adr_qd77.get_param_adr('JOGSPD_B_ADR',axis_no,qd77_axis_start)}")
        data.append(f"{motion.motor_param_adr_qd77.get_param_adr('U_START',axis_no)}\{motion.motor_param_adr_qd77.get_param_adr('JOGSPD_B_ADR',axis_no,qd77_axis_start)}")
        data.append(f"{motion.motor_param_adr_qd77.get_param_adr('U_START',axis_no)}\{motion.motor_param_adr_qd77.get_param_adr('J_SPDLIMIT_B_ADR',axis_no,qd77_axis_start)}")
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),1+scan_cnt))
        data.append("")
        data.append("F10")
        data.append("")
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+19+axis_address_offset),0))
        data.append("")
        data.append("F10")
        data.append("")
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+19+axis_address_offset),1))
        data.append("")
        data.append("F10")
        data.append("")
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+19+axis_address_offset),2))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),1+scan_cnt))
        data.append("")
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+1+axis_address_offset),0))
        data.append("F10")
        data.append("")
        data.append("")
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+axis_address_offset),3))
        data.append("")
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+19+axis_address_offset),3))
        data.append("")
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+0+axis_address_offset),4))
        data.append("")
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+19+axis_address_offset),4))
        data.append("")
        data.append(add_to_address('M600',axis_no-1))
        data.append("")
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+19+axis_address_offset),5))
        data.append(add_to_address(motion.motor_start_adr.get_start_adr('M_START',program_no),1+scan_cnt))
        data.append("")
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+19+axis_address_offset),0))
        data.append("K1")
        data.append(f"{motion.motor_param_adr_qd77.get_param_adr('U_START',axis_no)}\{motion.motor_param_adr_qd77.get_param_adr('JOG_FOR_B_ADR',axis_no,qd77_axis_start)}")
        data.append("")
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+19+axis_address_offset),1))
        data.append("K1")
        data.append(f"{motion.motor_param_adr_qd77.get_param_adr('U_START',axis_no)}\{motion.motor_param_adr_qd77.get_param_adr('JOG_REV_B_ADR',axis_no,qd77_axis_start)}")
        data.append("")
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+19+axis_address_offset),2))
        data.append("K1")
        data.append(f"{motion.motor_param_adr_qd77.get_param_adr('U_START',axis_no)}\{motion.motor_param_adr_qd77.get_param_adr('STOP_B_ADR',axis_no,qd77_axis_start)}")
        data.append("")
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+19+axis_address_offset),3))
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+19+axis_address_offset),4))
        data.append("")
        data.append("K1")
        data.append(f"{motion.motor_param_adr_qd77.get_param_adr('U_START',axis_no)}\{motion.motor_param_adr_qd77.get_param_adr('RESET_B_ADR',axis_no,qd77_axis_start)}")
        data.append("")    
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+19+axis_address_offset),5))
        data.append("K1")
        data.append(f"{motion.motor_param_adr_qd77.get_param_adr('U_START',axis_no)}\{motion.motor_param_adr_qd77.get_param_adr('SVOFF_B_ADR',axis_no,qd77_axis_start)}")
        data.append("")
        data.append(add_bit_offset(add_to_address(motion.motor_start_adr.get_start_adr('D_START',program_no),20+19+axis_address_offset),5))
        data.append("K0")
        data.append(f"{motion.motor_param_adr_qd77.get_param_adr('U_START',axis_no)}\{motion.motor_param_adr_qd77.get_param_adr('SVOFF_B_ADR',axis_no,qd77_axis_start)}")
        
        
    else:
        print("-")


# 3. D6부터 아래로 한 칸씩 채우기
start_row = 1
column_letter = 'E'
for i, value in enumerate(data):
    cell = f'{column_letter}{start_row + i}'  # D6, D7, D8...
    ws[cell] = value

cell = f'C{last_row}'
ws[cell] = 'END'

# 4. 저장

wb.save("motion_mecro_qd77_result2.xlsx")
wb.close()
print("✅ 저장 완료")




