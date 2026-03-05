import chardet
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
from mecro import MotionConfigurator,UnitConfigurator,add_to_address,add_to_bit_address,add_bit_offset
excel_file ='ULD_UNIT_MECRO_MAP.xlsx'
unit = UnitConfigurator(excel_file)




################################# 2. D6부터 아래로 채울 데이터################################################
#wb = Workbook()
#wb = load_workbook(excel_file)
#ws = wb['Debug']
wb= load_workbook("unit_mecro_result.xlsx")
ws = wb.active
ws.title = "Exported"
data = []
last_row = 2
program_no = 17

# B열 = 2번째 열, D열 = 4번째 열
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=5):
    for cell in row:
        if cell.column == 3 or cell.column == 5 or cell.column == 1:
            cell.value = None

##COMMON LADDER
last_row=unit.get_ladder_form.export_partial_to_excel(ws,last_row,"COMMON")
data.append("SM400")
data.append(unit.get_unit_adr.get_start_adr(program_no,"STATUS","M_START"))
data.append(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","M_START"))
data.append(unit.get_unit_adr.get_start_adr(program_no,"READY","M_START"))
data.append(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"))
data.append(unit.get_unit_adr.get_start_adr(program_no,"OUTPUT","M_START"))
data.append(unit.get_unit_adr.get_start_adr(program_no,"ERROR","M_START"))

cell = f'A{last_row}'
ws[cell] = '[Title]Unit Status Program'
last_row=last_row+1
data.append("")
cell = f'A{last_row}'
ws[cell] = 'Unit Cylinder Check'
last_row=last_row+1
data.append("")
##MANUAL LADDER_CYL

cyl_s_cnt=unit.get_unit_adr.get_refeat_count(program_no, "CYL_S_CNT")
for i in range(cyl_s_cnt):
    last_row=unit.get_ladder_form.export_partial_to_excel(ws,last_row,"STATUS_CYL")
    data.append(unit.get_unit_adr.get_start_adr(program_no,"STATUS","M_START"))
    data.append("")
    data.append("F13")
    data.append("F13")
    data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"STATUS","T_CYL"),i))
    data.append("K3")
    data.append("")
    data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"STATUS","T_CYL"),i))
    data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"STATUS","CYL_START"),i))
'''
cell = f'A{last_row}'
ws[cell] = 'Unit Vaccuume Check'
last_row=last_row+1
data.append("")
##MANUAL LADDER_VAC
vac_s_cnt=unit.get_unit_adr.get_refeat_count(program_no, "VAC_S_CNT")
#print(vac_s_cnt)
for i in range(vac_s_cnt):
    last_row=unit.get_ladder_form.export_partial_to_excel(ws,last_row,"STATUS_CYL")
    data.append(unit.get_unit_adr.get_start_adr(program_no,"STATUS","M_START"))
    data.append("")
    data.append("F13")
    data.append("F13")
    data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"STATUS","T_VAC"),i))
    data.append("K3")
    data.append("")
    data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"STATUS","T_VAC"),i))
    data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"STATUS","VAC_START"),i))
cell = f'A{last_row}'
ws[cell] = 'Unit Door Motor Stop'
last_row=last_row+1
data.append("")
##UNIT DOOR MOTOR STOP
last_row=unit.get_ladder_form.export_partial_to_excel(ws,last_row,"STATUS_DOOR")
data.append(unit.get_unit_adr.get_start_adr(program_no,"STATUS","M_START"))
data.append("F13")
data.append(add_to_address((unit.get_unit_adr.get_start_adr(program_no,"STATUS","M_START")),90))
data.append(unit.get_unit_adr.get_start_adr(program_no,"STATUS","M_START"))
data.append("F13")
data.append(add_to_address((unit.get_unit_adr.get_start_adr(program_no,"STATUS","M_START")),90))
data.append(add_to_address((unit.get_unit_adr.get_start_adr(program_no,"ERROR","M_START")),2))
data.append("M901")
data.append("")
data.append(add_to_address((unit.get_unit_adr.get_start_adr(program_no,"STATUS","M_START")),91))
'''
cell = f'A{last_row}'
ws[cell] = '[Title]Unit Manual Program'
last_row=last_row+1
data.append("")
cell = f'A{last_row}'
ws[cell] = 'Unit Manual Common'
last_row=last_row+1
data.append("")
##UNIT MANUAL_COMMON
last_row=unit.get_ladder_form.export_partial_to_excel(ws,last_row,"MANUAL_COMMON")
data.append(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","M_START"))
data.append("")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),5))
data.append("")
data.append(add_to_address((unit.get_unit_adr.get_start_adr(program_no,"STATUS","M_START")),290))
data.append(add_to_address((unit.get_unit_adr.get_start_adr(program_no,"ERROR","M_START")),5))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),13))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),18))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","M_START"),2))
data.append("")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),3))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","M_START"),3))
data.append("")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),12))
data.append(add_to_address((unit.get_unit_adr.get_start_adr(program_no,"ERROR","M_START")),5))
data.append("M10")
data.append("")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","M_START"),4))

cell = f'A{last_row}'
ws[cell] = 'Unit Cylinder Manual'
last_row=last_row+1
data.append("")
##UNIT CYL_MANUAL
cyl_m_cnt=unit.get_unit_adr.get_refeat_count(program_no, "CYL_M_CNT")
for i in range(cyl_m_cnt):
    last_row=unit.get_ladder_form.export_partial_to_excel(ws,last_row,"MANUAL_CYL")
    data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","M_START"),0))
    data.append("")
    data.append("F13")
    data.append("")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","CYL_START"),i),0))
    data.append("")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","CYL_START"),i),0))
    data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","M_START"),3))
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"OUTPUT","CYL_START"),i),1))
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","CYL_START"),i),1))
    data.append("")
    data.append("F13")
    data.append("")
    data.append("F13")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","CYL_START"),i),2))
    data.append("")
    data.append("F13")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","CYL_START"),i),3))
    data.append("")
    data.append("F13")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","CYL_START"),i),1))
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","CYL_START"),i),4))
'''
cell = f'A{last_row}'
ws[cell] = 'Unit Vaccuume Manual'
last_row=last_row+1
data.append("")
##UNIT VAC_MANUAL
vac_m_cnt=unit.get_unit_adr.get_refeat_count(program_no, "VAC_M_CNT")
for i in range(vac_m_cnt):
    last_row=unit.get_ladder_form.export_partial_to_excel(ws,last_row,"MANUAL_CYL")
    data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","M_START"),0))
    data.append("")
    data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"ERROR","VAC_START"),0+i))
    data.append("")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","VAC_START"),i),0))
    data.append("")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","VAC_START"),i),0))
    data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","M_START"),3))
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"OUTPUT","VAC_START"),i),1))
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","VAC_START"),i),1))
    data.append("")
    data.append("F13")
    data.append("")
    data.append("F13")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","VAC_START"),i),2))
    data.append("")
    data.append("F13")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","VAC_START"),i),3))
    data.append("")
    data.append("F13")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","VAC_START"),i),1))
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","VAC_START"),i),4))
'''
cell = f'A{last_row}'
ws[cell] = 'Unit SV Position Manual'
last_row=last_row+1
data.append("")
#UNIT MT_MANUAL
mt_m_cnt=unit.get_unit_adr.get_refeat_count(program_no, "MT_M_CNT")
for i in range(mt_m_cnt):
    last_row=unit.get_ladder_form.export_partial_to_excel(ws,last_row,"MANUAL_MT")
    data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","M_START"),0))
    data.append("")
    data.append("F13")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","MT_START"),i),0))
    data.append("")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","MT_START"),i),0))
    data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","M_START"),2))
    data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","M_START"),3))
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"OUTPUT","MT_START"),i),1))
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","MT_START"),i),1))
    data.append("")
    data.append("F13")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","MT_START"),i),2))
    data.append("")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","MT_START"),i),5))
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","MT_START"),i),3))
    data.append("")
    data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","M_START"),2))
    data.append("")
    data.append("F13")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","MT_START"),i),1))
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","MT_START"),i),5))
    data.append("")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","MT_START"),i),4))
    data.append("")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","MT_START"),i),4))
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","MT_START"),i),5))
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","MT_START"),i),4))
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","MT_START"),i),5))
    data.append("")
    data.append("")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","MT_START"),i),2))
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","MT_START"),i),0))
    data.append("M202")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","MT_START"),i),5))
'''
cell = f'A{last_row}'
ws[cell] = 'Unit Cycle Manual'
last_row=last_row+1
data.append("")
#UNIT CYCLE_MANUAL
cyc_m_cnt=unit.get_unit_adr.get_refeat_count(program_no, "CYC_M_CNT")
for i in range(cyc_m_cnt):
    last_row=unit.get_ladder_form.export_partial_to_excel(ws,last_row,"MANUAL_MT")
    data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","M_START"),0))
    data.append("")
    data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"ERROR","M_START"),5))
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","CYC_START"),i),0))
    data.append("")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","CYC_START"),i),0))
    data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","M_START"),2))
    data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","M_START"),3))
    data.append("F13")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","CYC_START"),i),1))
    data.append("")
    data.append("F13")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","CYC_START"),i),2))
    data.append("")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","CYC_START"),i),5))
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","CYC_START"),i),3))
    data.append("")
    data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","M_START"),2))
    data.append("")
    data.append("F13")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","CYC_START"),i),1))
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","CYC_START"),i),5))
    data.append("")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","CYC_START"),i),4))
    data.append("")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","CYC_START"),i),4))
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","CYC_START"),i),5))
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","CYC_START"),i),4))
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","CYC_START"),i),5))
    data.append("")
    data.append("")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","CYC_START"),i),2))
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","CYC_START"),i),0))
    data.append("M202")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","CYC_START"),i),5))
'''
cell = f'A{last_row}'
ws[cell] = '[Title]Unit Ready Program'
last_row=last_row+1
data.append("")
#UNIT_READY
last_row=unit.get_ladder_form.export_partial_to_excel(ws,last_row,"READY")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","M_START"),0))
data.append("")
data.append("M1155")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),3))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","M_START"),10))
data.append("")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","M_START"),13))
data.append("SM400")
data.append("")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"STATUS","M_START"),290))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),11))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"ERROR","M_START"),5))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),12))
data.append("M5")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","M_START"),14))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","M_START"),10))
data.append("")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","M_START"),10))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","M_START"),11))
data.append("")
data.append("F13")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","M_START"),13))
data.append("")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","M_START"),11))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","M_START"),24))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","M_START"),14))
data.append("")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","M_START"),13))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),12))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","M_START"),12))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","M_START"),14))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","M_START"),0))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","M_START"),11))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","D_START"),0))
data.append("K0")
data.append("K0")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","D_START"),0))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","M_START"),11))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","D_START"),0))
data.append("K0")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","M_START"),10))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","M_START"),20))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","D_START"),0))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","M_START"),11))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","D_START"),0))
data.append("K1")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","M_START"),21))
data.append("F13")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","D_START"),0))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","M_START"),11))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","D_START"),0))
data.append("K2")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","M_START"),22))
data.append("F13")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","D_START"),0))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","M_START"),11))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","D_START"),0))
data.append("K3")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","M_START"),23))
data.append("F13")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","D_START"),0))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","M_START"),11))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","D_START"),0))
data.append("K4")
data.append("F13")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","M_START"),24))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","D_START"),0))
cell = f'A{last_row}'
ws[cell] = '[Title]Unit Auto Program'
last_row=last_row+1
data.append("")
cell = f'A{last_row}'
ws[cell] = 'Unit Mode'
last_row=last_row+1
data.append("")
#UNIT AUTO PROGRAM
last_row=unit.get_ladder_form.export_partial_to_excel(ws,last_row,"AUTO_MODE")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),0))
data.append("")
data.append("M1160")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),13))
data.append("M1176")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),10))
data.append("")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","M_START"),10))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","M_START"),14))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),12))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),10))
data.append("")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),10))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),11))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),00))
data.append("")
data.append("M1165")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),12))
data.append("")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),13))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),12))
data.append("")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),18))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"READY","M_START"),20))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),20))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),13))
cell = f'A{last_row}'
ws[cell] = 'Unit Auto'
last_row=last_row+1
data.append("")
#UNIT AUTO
last_row=unit.get_ladder_form.export_partial_to_excel(ws,last_row,"UNIT_AUTO")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),0))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),11))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),15))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),0))
data.append("")
data.append("F13")
data.append("")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),16))
data.append("")
data.append("F13")
data.append("F13")
data.append("F13")
data.append("")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),17))
data.append("")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),15))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),16))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),17))
data.append("")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),18))
data.append("")
data.append("M1026")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),16))
data.append("")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),19))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),0))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),18))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","D_START"),0))
data.append("K0")
data.append("K0")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","D_START"),0))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),0))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","D_START"),0))
data.append("K0")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),20))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),18))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),19))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","D_START"),0))
data.append("K0")
data.append("F13")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","D_START"),0))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),18))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),19))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","D_START"),0))
data.append("K1")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),21))
data.append("F13")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","D_START"),0))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),18))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),19))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","D_START"),0))
data.append("K2")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),22))
data.append("F13")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","D_START"),0))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),18))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),19))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","D_START"),0))
data.append("K3")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","M_START"),23))
data.append("F13")
data.append("K0")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"AUTO","D_START"),0))
cell = f'A{last_row}'
ws[cell] = '[Title]Unit CYL/VAC Output Program'
last_row=last_row+1
data.append("")
#OUTPUT_CYL
cyl_o_cnt=unit.get_unit_adr.get_refeat_count(program_no, "CYL_O_CNT")
for i in range(cyl_o_cnt):
    last_row=unit.get_ladder_form.export_partial_to_excel(ws,last_row,"OUTPUT_CYL")
    data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"OUTPUT","M_START"),0))
    data.append("")
    data.append("F13")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"OUTPUT","CYL_START"),i),0))
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"OUTPUT","CYL_START"),i),1))
    data.append("")
    data.append("SM401")
    data.append("SM401")
    data.append("SM401")
    data.append("")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"OUTPUT","CYL_START"),i),0))
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","CYL_START"),i),4))
    data.append("SM401")
    data.append("SM401")
    data.append("")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"OUTPUT","CYL_START"),i),1))
    data.append("")
    data.append("")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"OUTPUT","CYL_START"),i),2))
'''
#OUTPUT_VAC
vac_o_cnt=unit.get_unit_adr.get_refeat_count(program_no, "VAC_O_CNT")
for i in range(vac_o_cnt):
    last_row=unit.get_ladder_form.export_partial_to_excel(ws,last_row,"OUTPUT_CYL")
    data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"OUTPUT","M_START"),0))
    data.append("")
    data.append("F13")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"OUTPUT","VAC_START"),i),0))
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"OUTPUT","VAC_START"),i),1))
    data.append("")
    data.append("SM401")
    data.append("SM401")
    data.append("SM401")
    data.append("")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"OUTPUT","VAC_START"),i),0))
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","VAC_START"),i),4))
    data.append("SM401")
    data.append("SM401")
    data.append("")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"OUTPUT","VAC_START"),i),1))
    data.append("")
    data.append("")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"OUTPUT","VAC_START"),i),2))
'''
cell = f'A{last_row}'
ws[cell] = '[Title]Unit MOTOR Output Program'
last_row=last_row+1
data.append("")
mt_o_cnt=unit.get_unit_adr.get_refeat_count(program_no, "MT_O_CNT")
for i in range(mt_o_cnt):
    last_row=unit.get_ladder_form.export_partial_to_excel(ws,last_row,"OUTPUT_MT")
    data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"OUTPUT","M_START"),0))
    data.append("")
    data.append("F13")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"OUTPUT","MT_START"),i),0))
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"OUTPUT","MT_START"),i),1))
    data.append("")
    data.append("SM401")
    data.append("SM401")
    data.append("SM401")
    data.append("")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"OUTPUT","MT_START"),i),0))
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"MANUAL","MT_START"),i),5))
    data.append("SM401")
    data.append("SM401")
    data.append("")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"OUTPUT","MT_START"),i),1))
    data.append("")
    data.append("")
    data.append(add_bit_offset(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"OUTPUT","MT_START"),i),2))
cell = f'A{last_row}'
'''
ws[cell] = '[Title]Unit Error Program'
last_row=last_row+1
data.append("")
cell = f'A{last_row}'
ws[cell] = 'Unit Error'
last_row=last_row+1
data.append("")
last_row=unit.get_ladder_form.export_partial_to_excel(ws,last_row,"ERROR_COMMON")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"ERROR","M_START"),0))
data.append("M5")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"ERROR","CYL_START"),0))
data.append("K90")
data.append("K0")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"ERROR","D_START"),0))
data.append("K20")
cell = f'A{last_row}'
ws[cell] = 'Cyl/Vac Error'
last_row=last_row+1
data.append("")
#ERROR_CYL
cyl_e_cnt=unit.get_unit_adr.get_refeat_count(program_no, "CYL_E_CNT")
for i in range(cyl_e_cnt):
    last_row=unit.get_ladder_form.export_partial_to_excel(ws,last_row,"ERROR_CYL")
    data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"ERROR","M_START"),0))
    data.append("")
    data.append("F13")
    data.append("F13")
    data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"ERROR","T_CYL"),i))
    data.append("K50")
    data.append("")
    data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"ERROR","T_CYL"),i))
    data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"ERROR","CYL_START"),i))
#ERROR_VAC
vac_e_cnt=unit.get_unit_adr.get_refeat_count(program_no, "VAC_E_CNT")
for i in range(vac_e_cnt):
    last_row=unit.get_ladder_form.export_partial_to_excel(ws,last_row,"ERROR_CYL")
    data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"ERROR","M_START"),0))
    data.append("")
    data.append("F13")
    data.append("F13")
    data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"ERROR","T_VAC"),i))
    data.append("K50")
    data.append("")
    data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"ERROR","T_VAC"),i))
    data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"ERROR","VAC_START"),i))
cell = f'A{last_row}'
ws[cell] = 'Motor Error'
last_row=last_row+1
data.append("")

#ERROR_MOTOR
last_row=unit.get_ladder_form.export_partial_to_excel(ws,last_row,"ERROR_MT")
cell = f'A{last_row-1}'
ws[cell] = 'Error Out'
data.append("")
data.append("")
#ERROR_OUT
last_row=unit.get_ladder_form.export_partial_to_excel(ws,last_row,"ERROR_OUT")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"ERROR","M_START"),0))
data.append("")
data.append("K0")
data.append(f'K4{add_to_address(unit.get_unit_adr.get_start_adr(program_no,"ERROR","CYL_START"),0)}')
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"ERROR","M_START"),1))
data.append("")
data.append("K0")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"ERROR","D_START"),0))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"ERROR","M_START"),2))
data.append("")
data.append("F13")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"ERROR","M_START"),3))
data.append("")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"ERROR","M_START"),1))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"ERROR","M_START"),2))
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"ERROR","M_START"),3))
data.append("")
data.append(add_to_address(unit.get_unit_adr.get_start_adr(program_no,"ERROR","M_START"),5))
'''
cell = f'C{last_row}'
ws[cell] = 'END'
# 3. D6부터 아래로 한 칸씩 채우기
#start_row = 1
#column_letter = 'B'
start_row = 2
column_letter = 'E'
for i, value in enumerate(data):
    cell = f'{column_letter}{start_row + i}'  # D6, D7, D8...
    ws[cell] = value

#cell = f'C{last_row}'
#ws[cell] = 'END'

# 4. 저장
#wb.save(excel_file)
wb.save("unit_mecro_result.xlsx")
print("✅ 저장 완료")




