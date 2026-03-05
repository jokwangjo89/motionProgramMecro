import pandas as pd
import re
from openpyxl import Workbook
import os

class MotionConfigurator:
    def __init__(self, excel_file):
        self.excel_file = excel_file

        # 각 서브 클래스 인스턴스 생성
        self.motor_prog = MotorProgLookup(excel_file)
        self.motor_pos_adr = MotorPosDataLookup(excel_file)
        self.motor_param_adr = MotorParamDataLookup(excel_file)
        self.motor_param_adr_qd75 = MotorParamDataLookup_QD75(excel_file)
        self.motor_param_adr_qd77 = MotorParamDataLookup_QD77(excel_file)
        self.motor_start_adr = MotorStartAdrLookup(excel_file)
        self.motor_motion_recv_adr = MotionAdrRecvLookup(excel_file)
        self.motor_motion_send_adr = MotionAdrSendLookup(excel_file)
        self.ladder_mecro = LadderMecro(excel_file)
        self.ladder_mecro_qd75 = LadderMecro_QD75(excel_file)
        self.ladder_mecro_qd77 = LadderMecro_QD77(excel_file)

class UnitConfigurator:
    def __init__(self,excel_file):
        self.excel_file = excel_file
        self.get_unit_adr= UnitStartAdrLookup(excel_file)
        self.get_ladder_form = UnitLadderMecro(excel_file)
    
class LadderMecro_QD75:
    def __init__(self, excel_file, sheet_name='LADDER_FORM_MOTOR_QD75'):
        # 행 인덱스로 첫 번째 열 사용 (START_ADR가 인덱스가 되도록)
        self.df = pd.read_excel(excel_file, sheet_name=sheet_name, engine='openpyxl', index_col=0)
        
    def export_to_excel(self, ws, start_row=1, start_col=1):
        self.df.columns = [
        col if not str(col).startswith("Unnamed") else f""
        for i, col in enumerate(self.df.columns)
        ]

        if self.df.index.name:
            ws.cell(row=start_row, column=start_col, value=self.df.index.name)
        else:
            ws.cell(row=start_row, column=start_col, value="")

        for j, col_name in enumerate(self.df.columns, start=start_col + 1):
            ws.cell(row=start_row, column=j, value=col_name)

        for i, (index_val, row) in enumerate(self.df.iterrows(), start=start_row + 1):
            ws.cell(row=i, column=start_col, value=index_val)
            for j, cell_value in enumerate(row, start=start_col + 1):
                ws.cell(row=i, column=j, value=cell_value)

        last_row = start_row + len(self.df)
        last_col = start_col + len(self.df.columns)
        print(f"📌 마지막 행: {last_row}, 마지막 열: {last_col}")
        return last_row + 1
    
    def export_partial_to_excel(self, ws, target_start_row, category):
        # Unnamed 컬럼 제거
        self.df.columns = [
            col if not str(col).startswith("Unnamed") else ""
            for col in self.df.columns
        ]

        source_col = 1
        target_col = 3

        # ✅ 열 개수 검사 (방어코드)
        if source_col >= len(self.df.columns):
            raise IndexError(f"source_col 인덱스 {source_col}가 유효하지 않습니다. 총 {len(self.df.columns)}개의 열만 존재합니다.")

        # ✅ 복사할 범위 지정
        source_start_row = 0
        source_end_row = len(self.df)
        if category == 'STATE':
            source_start_row = 1
            source_end_row = 36
        if category == 'ORIGIN':
            source_start_row = 37
            source_end_row = 75
        if category == 'POSITION_INTERLOCK':
            source_start_row = 76
            source_end_row = 123
        if category == 'POSITION_BIT':
            source_start_row = 124
            source_end_row = 265
        if category == 'POSITION_START':
            source_start_row = 266
            source_end_row = 350
        if category == 'POSITION_MOVEMENT':
            source_start_row = 351
            source_end_row = 556
        if category == 'POSITION_CHECK':
            source_start_row = 557
            source_end_row = 667
        if category == 'MANUAL_COMMAND':
            source_start_row = 668
            source_end_row = 743

        col_name = self.df.columns[source_col]

        # ✅ 데이터 자르기
        df_partial = self.df.loc[:, [col_name]].iloc[source_start_row:source_end_row]

        # 헤더 출력
        ws.cell(row=target_start_row, column=target_col, value=col_name)

        # ✅ 실제 데이터 출력 (예외 방지)
        for i, (_, row) in enumerate(df_partial.iterrows(), start=target_start_row):
            value = row[col_name]

            # Series나 리스트일 경우 첫 값만 사용
            if isinstance(value, (pd.Series, list)):
                value = value[0] if len(value) > 0 else ""

            # NaN 값은 공백으로 처리
            if pd.isna(value):
                value = ""

            # Excel에 출력
            ws.cell(row=i, column=target_col, value=value)

        last_row = target_start_row + len(df_partial)
        print(f"✅ 복사된 범위: 시트 기준 {target_start_row}행 ~ {last_row}행, 열 {target_col}")
        return last_row
class LadderMecro_QD77:
    def __init__(self, excel_file, sheet_name='LADDER_FORM_MOTOR_QD77'):
        # 행 인덱스로 첫 번째 열 사용 (START_ADR가 인덱스가 되도록)
        self.df = pd.read_excel(excel_file, sheet_name=sheet_name, engine='openpyxl', index_col=0)
        
    def export_to_excel(self, ws, start_row=1, start_col=1):
        self.df.columns = [
        col if not str(col).startswith("Unnamed") else f""
        for i, col in enumerate(self.df.columns)
        ]

        if self.df.index.name:
            ws.cell(row=start_row, column=start_col, value=self.df.index.name)
        else:
            ws.cell(row=start_row, column=start_col, value="")

        for j, col_name in enumerate(self.df.columns, start=start_col + 1):
            ws.cell(row=start_row, column=j, value=col_name)

        for i, (index_val, row) in enumerate(self.df.iterrows(), start=start_row + 1):
            ws.cell(row=i, column=start_col, value=index_val)
            for j, cell_value in enumerate(row, start=start_col + 1):
                ws.cell(row=i, column=j, value=cell_value)

        last_row = start_row + len(self.df)
        last_col = start_col + len(self.df.columns)
        print(f"📌 마지막 행: {last_row}, 마지막 열: {last_col}")
        return last_row + 1
    
    def export_partial_to_excel(self, ws, target_start_row, category):
        # Unnamed 컬럼 제거
        self.df.columns = [
            col if not str(col).startswith("Unnamed") else ""
            for col in self.df.columns
        ]

        source_col = 1
        target_col = 3

        # ✅ 열 개수 검사 (방어코드)
        if source_col >= len(self.df.columns):
            raise IndexError(f"source_col 인덱스 {source_col}가 유효하지 않습니다. 총 {len(self.df.columns)}개의 열만 존재합니다.")

        # ✅ 복사할 범위 지정
        source_start_row = 0
        source_end_row = len(self.df)
        if category == 'STATE':
            source_start_row = 1
            source_end_row = 34
        if category == 'ORIGIN':
            source_start_row = 35
            source_end_row = 73
        if category == 'POSITION_INTERLOCK':
            source_start_row = 74
            source_end_row = 121
        if category == 'POSITION_BIT':
            source_start_row = 122
            source_end_row = 263
        if category == 'POSITION_START':
            source_start_row = 264
            source_end_row = 348
        if category == 'POSITION_MOVEMENT':
            source_start_row = 349
            source_end_row = 554
        if category == 'POSITION_CHECK':
            source_start_row = 555
            source_end_row = 665
        if category == 'MANUAL_COMMAND':
            source_start_row = 666
            source_end_row = 747

        col_name = self.df.columns[source_col]

        # ✅ 데이터 자르기
        df_partial = self.df.loc[:, [col_name]].iloc[source_start_row:source_end_row]

        # 헤더 출력
        ws.cell(row=target_start_row, column=target_col, value=col_name)

        # ✅ 실제 데이터 출력 (예외 방지)
        for i, (_, row) in enumerate(df_partial.iterrows(), start=target_start_row):
            value = row[col_name]

            # Series나 리스트일 경우 첫 값만 사용
            if isinstance(value, (pd.Series, list)):
                value = value[0] if len(value) > 0 else ""

            # NaN 값은 공백으로 처리
            if pd.isna(value):
                value = ""

            # Excel에 출력
            ws.cell(row=i, column=target_col, value=value)

        last_row = target_start_row + len(df_partial)
        print(f"✅ 복사된 범위: 시트 기준 {target_start_row}행 ~ {last_row}행, 열 {target_col}")
        return last_row

class LadderMecro:
    def __init__(self, excel_file, sheet_name='LADDER_FORM_MOTOR'):
        # 행 인덱스로 첫 번째 열 사용 (START_ADR가 인덱스가 되도록)
        self.df = pd.read_excel(excel_file, sheet_name=sheet_name, engine='openpyxl', index_col=0)
        
    def export_to_excel(self, ws, start_row=1, start_col=1):
        self.df.columns = [
        col if not str(col).startswith("Unnamed") else f""
        for i, col in enumerate(self.df.columns)
        ]

        if self.df.index.name:
            ws.cell(row=start_row, column=start_col, value=self.df.index.name)
        else:
            ws.cell(row=start_row, column=start_col, value="")

        for j, col_name in enumerate(self.df.columns, start=start_col + 1):
            ws.cell(row=start_row, column=j, value=col_name)

        for i, (index_val, row) in enumerate(self.df.iterrows(), start=start_row + 1):
            ws.cell(row=i, column=start_col, value=index_val)
            for j, cell_value in enumerate(row, start=start_col + 1):
                ws.cell(row=i, column=j, value=cell_value)

        last_row = start_row + len(self.df)
        last_col = start_col + len(self.df.columns)
        print(f"📌 마지막 행: {last_row}, 마지막 열: {last_col}")
        return last_row + 1
    
    def export_partial_to_excel(self, ws, target_start_row, category):
        # Unnamed 컬럼 제거
        self.df.columns = [
            col if not str(col).startswith("Unnamed") else ""
            for col in self.df.columns
        ]

        source_col = 1
        target_col = 3

        # ✅ 열 개수 검사 (방어코드)
        if source_col >= len(self.df.columns):
            raise IndexError(f"source_col 인덱스 {source_col}가 유효하지 않습니다. 총 {len(self.df.columns)}개의 열만 존재합니다.")

        # ✅ 복사할 범위 지정
        source_start_row = 0
        source_end_row = len(self.df)
        if category == 'STATE':
            source_start_row = 1
            source_end_row = 40
        if category == 'ORIGIN':
            source_start_row = 41
            source_end_row = 74
        if category == 'POSITION_INTERLOCK':
            source_start_row = 75
            source_end_row = 117
        if category == 'POSITION_BIT':
            source_start_row = 118
            source_end_row = 264
        if category == 'POSITION_START':
            source_start_row = 265
            source_end_row = 346
        if category == 'POSITION_MOVEMENT':
            source_start_row = 347
            source_end_row = 543
        if category == 'POSITION_CHECK':
            source_start_row = 544
            source_end_row = 654
        if category == 'MANUAL_COMMAND':
            source_start_row = 655
            source_end_row = 740

        col_name = self.df.columns[source_col]

        # ✅ 데이터 자르기
        df_partial = self.df.loc[:, [col_name]].iloc[source_start_row:source_end_row]

        # 헤더 출력
        ws.cell(row=target_start_row, column=target_col, value=col_name)

        # ✅ 실제 데이터 출력 (예외 방지)
        for i, (_, row) in enumerate(df_partial.iterrows(), start=target_start_row):
            value = row[col_name]

            # Series나 리스트일 경우 첫 값만 사용
            if isinstance(value, (pd.Series, list)):
                value = value[0] if len(value) > 0 else ""

            # NaN 값은 공백으로 처리
            if pd.isna(value):
                value = ""

            # Excel에 출력
            ws.cell(row=i, column=target_col, value=value)

        last_row = target_start_row + len(df_partial)
        print(f"✅ 복사된 범위: 시트 기준 {target_start_row}행 ~ {last_row}행, 열 {target_col}")
        return last_row
class MotionAdrSendLookup:
    def __init__(self, excel_file, sheet_name='MOTIONCPU_ADR_PM'):
        # 행 인덱스로 첫 번째 열 사용 (START_ADR가 인덱스가 되도록)
        self.df = pd.read_excel(excel_file, sheet_name=sheet_name, engine='openpyxl', index_col=0)
        value = None
    def get_motion_axis_jogspeed_send_adr(self, param_name, axis_no, bit_no):
        if axis_no <= 32:
            description = 'M1_JOG_Speed'
        elif axis_no <= 64:
            description = 'M2_JOG_Speed'
            # bit_no = bit_no - 32
            axis_no = axis_no - 32
        else:
            return f"❌ 지원하지 않는 axis_no: {axis_no}"

        if param_name not in self.df.columns:
            return f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다."
        if description not in self.df.index:
            return f"❌ 행 '{description}' 이 존재하지 않습니다."

        device_type = self.get_type(description)
        base_value = self.df.loc[description, param_name]

        if pd.isna(base_value):
            return f"⚠️ '{param_name}' 값이 비어 있습니다."
        base_value = int(base_value + (axis_no-1)*self.get_length(description))

        if device_type == 'M':
            return f'{device_type}{base_value + bit_no}'

        elif device_type == 'D':
            if self.get_digit(description) == 1:
                return f'{device_type}{self.normalize_bit_address(base_value, bit_no)}'
            elif self.get_digit(description) == 16:
                return f'{device_type}{base_value}'
    def get_motion_axis_Command_send_adr(self, param_name, axis_no, bit_no):
        if axis_no <= 32:
            description = 'M1_Axis_Command'
        elif axis_no <= 64:
            description = 'M2_Axis_Command'
            # bit_no = bit_no - 32
            axis_no = axis_no - 32
        else:
            return f"❌ 지원하지 않는 axis_no: {axis_no}"

        if param_name not in self.df.columns:
            return f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다."
        if description not in self.df.index:
            return f"❌ 행 '{description}' 이 존재하지 않습니다."

        device_type = self.get_type(description)
        base_value = self.df.loc[description, param_name]

        if pd.isna(base_value):
            return f"⚠️ '{param_name}' 값이 비어 있습니다."
        base_value = int(base_value + (axis_no-1)*self.get_length(description))

        if device_type == 'M':
            return f'{device_type}{base_value + bit_no}'

        elif device_type == 'D':
            if self.get_digit(description) == 1:
                return f'{device_type}{self.normalize_bit_address(base_value, bit_no)}'
            elif self.get_digit(description) == 16:
                return f'{device_type}{base_value}'
    def get_motion_axis_position_send_adr(self, param_name, axis_no, bit_no):
        if axis_no <= 32:
            description = 'M1_Axis_Position_Address'
        elif axis_no <= 64:
            description = 'M2_Axis_Position_Address'
            # bit_no = bit_no - 32
            axis_no = axis_no - 32
        else:
            return f"❌ 지원하지 않는 axis_no: {axis_no}"

        if param_name not in self.df.columns:
            return f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다."
        if description not in self.df.index:
            return f"❌ 행 '{description}' 이 존재하지 않습니다."

        device_type = self.get_type(description)
        base_value = self.df.loc[description, param_name]

        if pd.isna(base_value):
            return f"⚠️ '{param_name}' 값이 비어 있습니다."
        base_value = int(base_value + (axis_no-1)*self.get_length(description))

        if device_type == 'M':
            return f'{device_type}{base_value + bit_no}'

        elif device_type == 'D':
            if self.get_digit(description) == 1:
                return f'{device_type}{self.normalize_bit_address(base_value, bit_no)}'
            elif self.get_digit(description) == 16:
                return f'{device_type}{base_value + bit_no}'
    def get_motion_axis_SFC_home_Send_adr(self, param_name, axis_no, bit_no):
        if axis_no <= 32:
            description = 'M1_SFC_Command_Home'
        elif axis_no <= 64:
            description = 'M2_SFC_Command_Home'
            bit_no = bit_no - 32
        else:
            return f"❌ 지원하지 않는 axis_no: {axis_no}"

        if param_name not in self.df.columns:
            return f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다."
        if description not in self.df.index:
            return f"❌ 행 '{description}' 이 존재하지 않습니다."

        device_type = self.get_type(description)
        base_value = self.df.loc[description, param_name]

        if pd.isna(base_value):
            return f"⚠️ '{param_name}' 값이 비어 있습니다."

        if device_type == 'M':
            return f'{device_type}{int(base_value + bit_no)}'

        elif device_type == 'D':
            return f'{device_type}{self.normalize_bit_address(int(base_value), bit_no-1)}'

        else:
            return f"⚠️ 지원하지 않는 device_type: {device_type}"       
    def get_motion_axis_SFC_pos_Send_adr(self, param_name, axis_no, bit_no):
        if axis_no <= 32:
            description = 'M1_SFC_Command_Pos'
        elif axis_no <= 64:
            description = 'M2_SFC_Command_Pos'
            bit_no = bit_no - 32
        else:
            return f"❌ 지원하지 않는 axis_no: {axis_no}"

        if param_name not in self.df.columns:
            return f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다."
        if description not in self.df.index:
            return f"❌ 행 '{description}' 이 존재하지 않습니다."

        device_type = self.get_type(description)
        base_value = self.df.loc[description, param_name]

        if pd.isna(base_value):
            return f"⚠️ '{param_name}' 값이 비어 있습니다."

        if device_type == 'M':
            return f'{device_type}{int(base_value + bit_no)}'

        elif device_type == 'D':
            return f'{device_type}{self.normalize_bit_address(int(base_value), bit_no-1)}'

        else:
            return f"⚠️ 지원하지 않는 device_type: {device_type}"       
    def get_type(self, description, param_name='TYPE'):
        if param_name not in self.df.columns:
            return f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다."
        if description not in self.df.index:
            return f"❌ 행 '{description}' 이 존재하지 않습니다."

        value = self.df.loc[description, param_name]
        return value if pd.notna(value) else f"⚠️ '{param_name}' 값이 비어 있습니다."
    def get_digit(self, description, param_name='DIGIT'):
        if param_name not in self.df.columns:
            return f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다."
        if description not in self.df.index:
            return f"❌ 행 '{description}' 이 존재하지 않습니다."

        value = self.df.loc[description, param_name]
        return value if pd.notna(value) else f"⚠️ '{param_name}' 값이 비어 있습니다."
    def get_length(self, description, param_name='LENGTH'):
        if param_name not in self.df.columns:
            return f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다."
        if description not in self.df.index:
            return f"❌ 행 '{description}' 이 존재하지 않습니다."

        value = self.df.loc[description, param_name]
        return value if pd.notna(value) else f"⚠️ '{param_name}' 값이 비어 있습니다."  
    def to_hex_char(self, n: int) -> str:
        return format(n, 'X')  # 또는 'x' 소문자
    def normalize_bit_address(self, base_addr: int, bit_no: int) -> str:
        word_offset = (bit_no) // 16
        bit_offset = (bit_no) % 16
        return f"{base_addr + word_offset}.{self.to_hex_char(bit_offset)}"
class MotionAdrRecvLookup:
    def __init__(self, excel_file, sheet_name='MOTIONCPU_ADR_MP'):
        # 행 인덱스로 첫 번째 열 사용 (START_ADR가 인덱스가 되도록)
        self.df = pd.read_excel(excel_file, sheet_name=sheet_name, engine='openpyxl', index_col=0)
        value = None
    def get_motion_axis_status_recv_adr(self, param_name,axis_no,bit_no):
        if axis_no <= 32:
            description = 'M1_Axis_Status'
        elif axis_no <= 64:
            description = 'M2_Axis_Status'
            # bit_no = bit_no - 32
            axis_no = axis_no - 32
        else:
            return f"❌ 지원하지 않는 axis_no: {axis_no}"

        if param_name not in self.df.columns:
            return f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다."
        if description not in self.df.index:
            return f"❌ 행 '{description}' 이 존재하지 않습니다."

        device_type = self.get_type(description)
        base_value = self.df.loc[description, param_name]

        if pd.isna(base_value):
            return f"⚠️ '{param_name}' 값이 비어 있습니다."
        base_value = int(base_value + (axis_no-1)*self.get_length(description)+ bit_no)

        if device_type == 'M':
            return f'{device_type}{base_value}'

        elif device_type == 'D':
            if self.get_digit(description) == 1:
                return f'{device_type}{self.normalize_bit_address(base_value, bit_no)}'
            elif self.get_digit(description) == 16:
                return f'{device_type}{base_value}'
    def get_motion_axis_monitor_recv_adr(self, param_name,axis_no,bit_no):
        if axis_no <= 32:
            description = 'M1_Axis_Monitor_Devices'
        elif axis_no <= 64:
            description = 'M2_Axis_Monitor_Devices'
            # bit_no = bit_no - 32
            axis_no = axis_no - 32
        else:
            return f"❌ 지원하지 않는 axis_no: {axis_no}"

        if param_name not in self.df.columns:
            return f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다."
        if description not in self.df.index:
            return f"❌ 행 '{description}' 이 존재하지 않습니다."

        device_type = self.get_type(description)
        base_value = self.df.loc[description, param_name]

        if pd.isna(base_value):
            return f"⚠️ '{param_name}' 값이 비어 있습니다."
        base_value = int(base_value + (axis_no-1)*self.get_length(description))

        if device_type == 'M':
            return f'{device_type}{base_value}'

        elif device_type == 'D':
            if self.get_digit(description) == 1:
                return f'{device_type}{self.normalize_bit_address(base_value, bit_no)}'
            elif self.get_digit(description) == 16:
                return f'{device_type}{base_value+ bit_no}'
    def get_motion_axis_SFC_home_recv_adr(self, param_name, axis_no, bit_no):
        if axis_no <= 32:
            description = 'M1_SFC_Monitor_HomeBusy'
        elif axis_no <= 64:
            description = 'M2_SFC_Monitor_HomeBusy'
            bit_no = bit_no - 32
        else:
            return f"❌ 지원하지 않는 axis_no: {axis_no}"

        if param_name not in self.df.columns:
            return f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다."
        if description not in self.df.index:
            return f"❌ 행 '{description}' 이 존재하지 않습니다."

        device_type = self.get_type(description)
        base_value = self.df.loc[description, param_name]

        if pd.isna(base_value):
            return f"⚠️ '{param_name}' 값이 비어 있습니다."

        if device_type == 'M':
            return f'{device_type}{int(base_value + bit_no)}'

        elif device_type == 'D':
            return f'{device_type}{self.normalize_bit_address(int(base_value), bit_no-1)}'

        else:
            return f"⚠️ 지원하지 않는 device_type: {device_type}"
    def get_motion_axis_SFC_pos_recv_adr(self, param_name, axis_no, bit_no):
        if axis_no <= 32:
            description = 'M1_SFC_Monitor_PosBusy'
        elif axis_no <= 64:
            description = 'M2_SFC_Monitor_PosBusy'
            bit_no = bit_no - 32
        else:
            return f"❌ 지원하지 않는 axis_no: {axis_no}"

        if param_name not in self.df.columns:
            return f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다."
        if description not in self.df.index:
            return f"❌ 행 '{description}' 이 존재하지 않습니다."

        device_type = self.get_type(description)
        base_value = self.df.loc[description, param_name]

        if pd.isna(base_value):
            return f"⚠️ '{param_name}' 값이 비어 있습니다."

        if device_type == 'M':
            return f'{device_type}{int(base_value + bit_no)}'

        elif device_type == 'D':
            return f'{device_type}{self.normalize_bit_address(int(base_value), bit_no-1)}'

        else:
            return f"⚠️ 지원하지 않는 device_type: {device_type}"
    def get_motion_axis_SFC_axis_recv_adr(self, param_name, axis_no, bit_no):
        if axis_no <= 32:
            description = 'M1_SFC_Monitor_AxisBusy'
        elif axis_no <= 64:
            description = 'M2_SFC_Monitor_AxisBusy'
            bit_no = bit_no - 32
        else:
            return f"❌ 지원하지 않는 axis_no: {axis_no}"

        if param_name not in self.df.columns:
            return f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다."
        if description not in self.df.index:
            return f"❌ 행 '{description}' 이 존재하지 않습니다."

        device_type = self.get_type(description)
        base_value = self.df.loc[description, param_name]

        if pd.isna(base_value):
            return f"⚠️ '{param_name}' 값이 비어 있습니다."

        if device_type == 'M':
            return f'{device_type}{int(base_value + bit_no)}'

        elif device_type == 'D':
            return f'{device_type}{self.normalize_bit_address(int(base_value), bit_no-1)}'

        else:
            return f"⚠️ 지원하지 않는 device_type: {device_type}"
                          
    def get_type(self, description, param_name='TYPE'):
        if param_name not in self.df.columns:
            return f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다."
        if description not in self.df.index:
            return f"❌ 행 '{description}' 이 존재하지 않습니다."

        value = self.df.loc[description, param_name]
        return value if pd.notna(value) else f"⚠️ '{param_name}' 값이 비어 있습니다."
    def get_digit(self, description, param_name='DIGIT'):
        if param_name not in self.df.columns:
            return f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다."
        if description not in self.df.index:
            return f"❌ 행 '{description}' 이 존재하지 않습니다."

        value = self.df.loc[description, param_name]
        return value if pd.notna(value) else f"⚠️ '{param_name}' 값이 비어 있습니다."
    def get_length(self, description, param_name='LENGTH'):
        if param_name not in self.df.columns:
            return f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다."
        if description not in self.df.index:
            return f"❌ 행 '{description}' 이 존재하지 않습니다."

        value = self.df.loc[description, param_name]
        return value if pd.notna(value) else f"⚠️ '{param_name}' 값이 비어 있습니다."  
    def to_hex_char(self, n: int) -> str:
        return format(n, 'X')  # 또는 'x' 소문자
    def normalize_bit_address(self, base_addr: int, bit_no: int) -> str:
        word_offset = (bit_no) // 16
        bit_offset = (bit_no) % 16
        return f"{base_addr + word_offset}.{self.to_hex_char(bit_offset)}"
class MotorStartAdrLookup:
    def __init__(self, excel_file, sheet_name='PROG_ADR'):
        # 행 인덱스로 첫 번째 열 사용 (START_ADR가 인덱스가 되도록)
        self.df = pd.read_excel(excel_file, sheet_name=sheet_name, engine='openpyxl', index_col=0)
        value = None

    def get_start_adr(self, param_name,program_no):  

        if param_name not in self.df.columns:
            return f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다."
        if program_no not in self.df.index:
            return f"❌ 행 '{program_no}' 이 존재하지 않습니다."
        device_type = param_name[0]
        value = self.df.loc[program_no,param_name]
        if pd.notna(value):
            return f'{device_type}{int(value)}'
        else:
            return f"⚠️ '{param_name}' 값이 비어 있습니다."
class MotorParamDataLookup:
    def __init__(self, excel_file, sheet_name='MOTOR_PARA_ADR'):
        # 행 인덱스로 첫 번째 열 사용 (START_ADR가 인덱스가 되도록)
        self.df = pd.read_excel(excel_file, sheet_name=sheet_name, engine='openpyxl', index_col=0)
        value = None
        self.row_name = 'START_ADR'

    def get_param_adr(self, param_name,axis_no):  

        if param_name not in self.df.columns:
            return f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다."
        if self.row_name not in self.df.index:
            return f"❌ 행 '{self.row_name}' 이 존재하지 않습니다."

        value = self.df.loc[self.row_name, param_name]
        value = (axis_no-1) * self.get_pos_length() + value
        if pd.notna(value):
            return f'{self.get_pos_type()}{int(value)}'
        else:
            return f"⚠️ '{param_name}' 값이 비어 있습니다."
    def get_pos_length(self,param_name='AXIS_LENGTH'):
        if param_name not in self.df.columns:
            return f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다."
        if self.row_name not in self.df.index:
            return f"❌ 행 '{self.row_name}' 이 존재하지 않습니다."

        value = self.df.loc[self.row_name, param_name]
        return int(value) if pd.notna(value) else f"⚠️ '{param_name}' 값이 비어 있습니다."
    def get_pos_type(self,param_name='TYPE'):
        if param_name not in self.df.columns:
            return f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다."
        if self.row_name not in self.df.index:
            return f"❌ 행 '{self.row_name}' 이 존재하지 않습니다."

        value = self.df.loc[self.row_name, param_name]
        return value if pd.notna(value) else f"⚠️ '{param_name}' 값이 비어 있습니다."

class MotorParamDataLookup_QD75:
    def __init__(self, excel_file, sheet_name='QD75_PARA_ADR'):
        self.df = pd.read_excel(excel_file, sheet_name=sheet_name, engine='openpyxl', index_col=0)
        self.df.columns = self.df.columns.str.strip()
        self.df.index = self.df.index.astype(int)
        self.row_name = None

    def get_param_adr(self, param_name, axis_no, start_point=None):  
        if start_point is None:
            start_point = 0
        self.row_name = int(axis_no)

        # param_name별 설정
        param_info = {
            "CURRENT_POS_ADR": ("CURRENT_POS_TYPE", "CURRENT_POS_LEN"),
            "EXT_IO_ADR": ("EXT_IO_TYPE", "EXT_IO_LEN"),
            "START_X_ADR": ("START_X_TYPE", "START_X_LEN"),
            "START_Y_ADR": ("START_Y_TYPE", "START_Y_LEN"),
            "U_START": ("U_TYPE", "U_LEN"),
            "POSNO_B_ADR": ("B_TYPE", "POSNO_B_LEN"),
            "MCODE_B_ADR": ("B_TYPE", "MCODE_B_LEN"),
            "ACC_B_ADR": ("B_TYPE", "ACC_B_LEN"),
            "DEC_B_ADR": ("B_TYPE", "DEC_B_LEN"),
            "SPEED_B_ADR": ("B_TYPE", "SPEED_B_LEN"),
            "A_SPDLIMIT_B_ADR": ("B_TYPE", "A_SPDLIMIT_B_LEN"),
            "POSADR_B_ADR": ("B_TYPE", "POSADR_B_LEN"),
            "J_SPDLIMIT_B_ADR": ("B_TYPE", "J_SPDLIMIT_B_LEN"),
            "JOGSPD_B_ADR": ("B_TYPE", "JOGSPD_B_LEN"),
            "JOG_Y_ADR": ("JOG_Y_TYPE", "JOG_Y_LEN"),
            "RESET_B_ADR": ("B_TYPE", "RESET_B_LEN"),
            "SVOFF_B_ADR": ("B_TYPE", "SVOFF_B_LEN"),
        }

        if param_name not in self.df.columns:
            return f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다."
        if self.row_name not in self.df.index:
            return f"❌ 행 '{self.row_name}' 이 존재하지 않습니다."
        if param_name not in param_info:
            return f"❌ '{param_name}' 이 정의되어 있지 않습니다."

        param_type_col, param_len_col = param_info[param_name]
        raw_value = self.df.loc[self.row_name, param_name]

        # DEBUG LOG
        print("💬 START_X_ADR 값:", raw_value, "| 타입:", type(raw_value))

        # U_START는 보정 없이 그대로 반환
        if param_name == "U_START":
            return f'{self.get_axis_type(param_type_col).strip().upper()}{raw_value}'
        

        # 주소 계산
        length = self.get_axis_length(param_len_col)
        
        offset = (axis_no - start_point - 1) * length

        if pd.notna(raw_value):
            raw_str = str(raw_value).strip()
            type_prefix = self.get_axis_type(param_type_col).strip().upper()
            print("📌 type_prefix:", type_prefix)

            try:
                # 순수 숫자형만 인식
                if type_prefix in ["X", "Y"]:
                    match = re.match(r"^([A-Fa-f0-9]+)$", raw_str)
                    if not match:
                        return f"❌ 16진수 주소 형식 오류: '{raw_str}'"
                    base_hex = match.group(1)
                    base = int(base_hex, 16)
                    result = base + offset
                    return f"{type_prefix}{format(result, 'X').upper()}"
                else:
                    base = int(raw_str)
                    result = base + offset
                    return f"{type_prefix}{result}"

            except ValueError:
                return f"❌ '{raw_str}' 은(는) 주소로 변환할 수 없습니다."
        else:
            return f"⚠️ '{param_name}' 값이 비어 있습니다."

    def get_axis_length(self, param_name):
        if param_name not in self.df.columns:
            raise ValueError(f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다.")
        value = self.df.loc[self.row_name, param_name]
        return int(value) if pd.notna(value) else 0

    def get_axis_type(self, param_name):
        if param_name not in self.df.columns:
            return f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다."
        value = self.df.loc[self.row_name, param_name]
        return value.strip() if pd.notna(value) else ""
    def get_start_axis(self,param_name, axis_no):
        if param_name not in self.df.columns:
            return f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다."
        if axis_no not in self.df.index:
            return f"❌ 행 '{axis_no}' 이 존재하지 않습니다."

        raw_value = self.df.loc[axis_no, param_name]
        return raw_value
class MotorParamDataLookup_QD77:
    def __init__(self, excel_file, sheet_name='QD77_PARA_ADR'):
        self.df = pd.read_excel(excel_file, sheet_name=sheet_name, engine='openpyxl', index_col=0)
        self.df.columns = self.df.columns.str.strip()
        self.df = self.df[self.df.index.notna()]
        self.df.index = self.df.index.astype(int)
        self.row_name = None

    def get_param_adr(self, param_name, axis_no, start_point=None):  
        if start_point is None:
            start_point = 0
        self.row_name = int(axis_no)

        # param_name별 설정
        param_info = {
            "CURRENT_POS_ADR": ("CURRENT_POS_TYPE", "CURRENT_POS_LEN"),
            "EXT_IO_ADR": ("EXT_IO_TYPE", "EXT_IO_LEN"),
            "START_X_ADR": ("START_X_TYPE", "START_X_LEN"),
            "START_Y_ADR": ("START_Y_TYPE", "START_Y_LEN"),
            "U_START": ("U_TYPE", "U_LEN"),
            "POSNO_B_ADR": ("B_TYPE", "POSNO_B_LEN"),
            "MCODE_B_ADR": ("B_TYPE", "MCODE_B_LEN"),
            "ACC_B_ADR": ("B_TYPE", "ACC_B_LEN"),
            "DEC_B_ADR": ("B_TYPE", "DEC_B_LEN"),
            "SPEED_B_ADR": ("B_TYPE", "SPEED_B_LEN"),
            "A_SPDLIMIT_B_ADR": ("B_TYPE", "A_SPDLIMIT_B_LEN"),
            "POSADR_B_ADR": ("B_TYPE", "POSADR_B_LEN"),
            "J_SPDLIMIT_B_ADR": ("B_TYPE", "J_SPDLIMIT_B_LEN"),
            "JOGSPD_B_ADR": ("B_TYPE", "JOGSPD_B_LEN"),
            "JOG_Y_ADR": ("JOG_Y_TYPE", "JOG_Y_LEN"),
            "RESET_B_ADR": ("B_TYPE", "RESET_B_LEN"),
            "SVOFF_B_ADR": ("B_TYPE", "SVOFF_B_LEN"),
            "STATUS_ADR": ("STATUS_TYPE", "STATUS_LEN"),
            "JOG_FOR_B_ADR": ("B_TYPE", "JOG_FOR_B_LEN"),
            "JOG_REV_B_ADR": ("B_TYPE", "JOG_REV_B_LEN"),
            "STOP_B_ADR": ("B_TYPE", "JOG_REV_B_LEN"),

        }

        if param_name not in self.df.columns:
            return f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다."
        if self.row_name not in self.df.index:
            return f"❌ 행 '{self.row_name}' 이 존재하지 않습니다."
        if param_name not in param_info:
            return f"❌ '{param_name}' 이 정의되어 있지 않습니다."

        param_type_col, param_len_col = param_info[param_name]
        raw_value = self.df.loc[self.row_name, param_name]

        # U_START는 보정 없이 그대로 반환
        if param_name == "U_START":
            return f'{self.get_axis_type(param_type_col).strip().upper()}{raw_value}'

        # 주소 계산
        length = self.get_axis_length(param_len_col)
        offset = (axis_no - start_point - 1) * length

        if isinstance(raw_value, pd.Series):
            raw_value = raw_value.iloc[0]

        if pd.notna(raw_value):
            type_prefix = self.get_axis_type(param_type_col).strip().upper()
            print("💬 값:", raw_value, "| 타입:", type(raw_value), "| 접두사:", type_prefix)

            try:
                raw_str = str(raw_value).strip()

                if type_prefix in ["X", "Y"]:
                    # 16진수 처리
                    match = re.match(r"^([A-Fa-f0-9]+)$", raw_str)
                    if not match:
                        return f"❌ 16진수 주소 형식 오류: '{raw_str}'"
                    base = int(raw_str, 16)
                    result = base + offset
                    return f"{type_prefix}{format(result, 'X').upper()}"
                else:
                    # 10진수 처리 (문자든 숫자든 float 변환 후 int)
                    base = int(float(raw_str))
                    result = base + offset
                    return f"{type_prefix}{result}"

            except Exception as e:
                return f"❌ 주소 변환 실패: {raw_value} ({e})"
        else:
            return f"⚠️ '{param_name}' 값이 비어 있습니다."

    def get_axis_length(self, param_name):
        if param_name not in self.df.columns:
            raise ValueError(f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다.")
        value = self.df.loc[self.row_name, param_name]

        # Series일 경우 첫 번째 값만 사용
        if isinstance(value, pd.Series):
            value = value.iloc[0]
            

        return int(value) if pd.notna(value) else 0

    def get_axis_type(self, param_name):
        if param_name not in self.df.columns:
            return f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다."
        value = self.df.loc[self.row_name, param_name]
        return value.strip() if pd.notna(value) else ""
    def get_start_axis(self, param_name, axis_no):
        axis_no = int(axis_no)  # ← 혹시 float으로 넘어올 수 있으니 강제 정수화
        print(param_name, axis_no)

        if param_name not in self.df.columns:
            return f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다."
        if axis_no not in self.df.index:
            print("❌ 존재하는 인덱스 목록:", self.df.index.tolist())  # 디버깅용 출력
            return f"❌ 행 '{axis_no}' 이 존재하지 않습니다."

        return self.df.loc[axis_no, param_name]

class MotorPosDataLookup:
    def __init__(self, excel_file, sheet_name='MOTOR_POS_ADR'):
        # 행 인덱스로 첫 번째 열 사용 (START_ADR가 인덱스가 되도록)
        self.df = pd.read_excel(excel_file, sheet_name=sheet_name, engine='openpyxl', index_col=0)
        value = None
        self.row_name = 'START_ADR'

    def get_pos_adr(self, param_name,axis_no,pos_no):
        if param_name not in self.df.columns:
            return f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다."
        if self.row_name not in self.df.index:
            return f"❌ 행 '{self.row_name}' 이 존재하지 않습니다."

        value = self.df.loc[self.row_name, param_name]
        value = (axis_no-1) * self.get_axis_length() + value + (pos_no - 1) * self.get_pos_length()
        if pd.notna(value):
            return f'{self.get_pos_type()}{int(value)}'
        else:
            return f"⚠️ '{param_name}' 값이 비어 있습니다."
    def get_pos_length(self,param_name='POS_LENGTH'):
        if param_name not in self.df.columns:
            return f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다."
        if self.row_name not in self.df.index:
            return f"❌ 행 '{self.row_name}' 이 존재하지 않습니다."

        value = self.df.loc[self.row_name, param_name]
        return int(value) if pd.notna(value) else f"⚠️ '{param_name}' 값이 비어 있습니다."
    def get_axis_length(self,param_name='AXIS_LENGTH'):
        if param_name not in self.df.columns:
            return f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다."
        if self.row_name not in self.df.index:
            return f"❌ 행 '{self.row_name}' 이 존재하지 않습니다."

        value = self.df.loc[self.row_name, param_name]
        return int(value) if pd.notna(value) else f"⚠️ '{param_name}' 값이 비어 있습니다."
    def get_pos_type(self,param_name='TYPE'):
        if param_name not in self.df.columns:
            return f"❌ 컬럼 '{param_name}' 이 존재하지 않습니다."
        if self.row_name not in self.df.index:
            return f"❌ 행 '{self.row_name}' 이 존재하지 않습니다."

        value = self.df.loc[self.row_name, param_name]
        return value if pd.notna(value) else f"⚠️ '{param_name}' 값이 비어 있습니다."
class MotorProgLookup:
    def __init__(self, excel_file, sheet_name='MOTOR_PROG_NO'):
        self.excel_file = excel_file
        self.sheet_name = sheet_name
        self.df = pd.read_excel(excel_file, sheet_name=sheet_name, engine='openpyxl')

    def get_axis_no(self, program_no, column_name):
        if column_name not in self.df.columns:
            return f"❌ 컬럼 '{column_name}' 이 존재하지 않습니다."
        
        row = self.df[self.df['Program No'] == program_no]

        if row.empty:
            return f"❌ Program No '{program_no}' 이 존재하지 않습니다."

        value = row.iloc[0][column_name]

        return int(value) if pd.notna(value) else f"⚠️ '{column_name}' 값이 비어 있습니다."

    def print_data(self):
        print(self.df)

class UnitStartAdrLookup:
    def __init__(self, excel_file, sheet_name='PROG_ADR'):
        self.excel_file = excel_file
        self.df = pd.read_excel(excel_file, sheet_name=sheet_name, engine='openpyxl', index_col=0)
        self.df.columns = [str(col).strip() for col in self.df.columns]
    def get_start_adr(self, raw_index, cat, unit):
        # ▶ cat/unit 이 있다면 그 값을 먼저 출력해봄
        if cat is not None and unit is not None:
            result = self.get_prog_1unit_adr(cat, unit)

            if isinstance(result, str):
                # 에러 메시지라면 그대로 리턴
                return result

            adr_offs, adr_type = result
            
            #print(f"🔎 [참조 값] {unit} of {cat} → {adr_offs}")
            type = adr_type
        # type이 정상적으로 지정되지 않은 경우 방어
        if 'type' not in locals():
            return "❌ type 값이 설정되지 않았습니다."

        # 타입별 시작주소 컬럼 지정
        if type == 'M':
            columns = 'M_START'
        elif type == 'D':
            columns = 'D_START'
        elif type == 'T':
            columns = 'T_START'
        else:
            return f"❌ 알 수 없는 type '{type}'"

        # 컬럼 및 행 존재 여부 체크
        if columns not in self.df.columns:
            return f"❌ 컬럼 '{columns}' 이 존재하지 않습니다."
        if raw_index not in self.df.index:
            return f"❌ 행 '{raw_index}' 이 존재하지 않습니다."

        base_value = self.df.loc[raw_index, columns]

        if pd.isna(base_value):
            return f"⚠️ '{columns}' 값이 비어 있습니다."

        base_value = int(base_value) + int(adr_offs)
        return f'{type}{base_value}'
    def get_prog_1unit_adr(self, raw_index, columns):
        df = pd.read_excel(self.excel_file, sheet_name='UNIT_ADR', engine='openpyxl', index_col=0)

        if columns not in df.columns:
            return f"❌ 컬럼 '{columns}' 이 존재하지 않습니다."
        if raw_index not in df.index:
            return f"❌ 행 '{raw_index}' 이 존재하지 않습니다."

        base_value = df.loc[raw_index, columns]

        if columns == "CYL_START":
             base_type = df.loc[raw_index, "CYL_TYPE"]
        elif columns == "VAC_START":
             base_type = df.loc[raw_index, "VAC_TYPE"]
        elif columns == "MT_START":
             base_type = df.loc[raw_index, "MT_TYPE"]
        elif columns == "CYC_START":
             base_type = df.loc[raw_index, "CYC_TYPE"]
        elif columns == "M_START":
             base_type = "M"
        elif columns == "D_START":
             base_type = "D"
        elif columns == "T_START":
             base_type = "T"
        elif columns == "T_CYL":
             base_type = "T"
        elif columns == "T_VAC":
             base_type = "T"
        if pd.isna(base_value):
            return f"⚠️ '{columns}' 값이 비어 있습니다."
        return base_value, base_type

    def get_refeat_count(self, raw_index, columns):
         # 컬럼 및 행 존재 여부 체크
        if columns not in self.df.columns:
            return f"❌ 컬럼 '{columns}' 이 존재하지 않습니다."
        if raw_index not in self.df.index:
            return f"❌ 행 '{raw_index}' 이 존재하지 않습니다."
        base_value = self.df.loc[raw_index, columns]
        if pd.isna(base_value):
            return f"⚠️ '{columns}' 값이 비어 있습니다."
        return int(base_value)
    
class UnitLadderMecro:
    def __init__(self, excel_file, sheet_name='LADDER_FORM'):
        # 첫 번째 열을 인덱스로 사용
        self.df = pd.read_excel(excel_file, sheet_name=sheet_name, engine='openpyxl', index_col=0)

    def export_partial_to_excel(self, ws, target_start_row, category):
        # Unnamed 컬럼 제거
        self.df.columns = [
            col if not str(col).startswith("Unnamed") else ""
            for col in self.df.columns
        ]

        source_col = 1
        target_col = 3

        # ✅ 열 개수 검사 (방어코드)
        if source_col >= len(self.df.columns):
            raise IndexError(f"source_col 인덱스 {source_col}가 유효하지 않습니다. 총 {len(self.df.columns)}개의 열만 존재합니다.")

        # ✅ 복사할 범위 지정
        source_start_row = 0
        source_end_row = len(self.df)
        if category == 'COMMON':
            source_start_row = 0
            source_end_row = 7
        if category == 'STATUS_CYL':
            source_start_row = 9
            source_end_row = 18
        if category == 'STATUS_DOOR':
            source_start_row = 119
            source_end_row = 129
        if category == 'MANUAL_COMMON':
            source_start_row = 131
            source_end_row = 149
        if category == 'MANUAL_CYL':
            source_start_row = 150
            source_end_row = 172
        if category == 'MANUAL_MT':
            source_start_row = 416
            source_end_row = 451
        if category == 'MANUAL_CYC':
            source_start_row = 522
            source_end_row = 557
        if category == 'READY':
            source_start_row = 628
            source_end_row = 695
        if category == 'AUTO_MODE':
            source_start_row = 697
            source_end_row = 723
        if category == 'UNIT_AUTO':
            source_start_row = 724
            source_end_row = 787
        if category == 'OUTPUT_CYL':
            source_start_row = 788
            source_end_row = 807
        if category == 'OUTPUT_MT':
            source_start_row = 1017
            source_end_row = 1036
        if category == 'ERROR_COMMON':
            source_start_row = 1076
            source_end_row = 1083
        if category == 'ERROR_CYL':
            source_start_row = 1084
            source_end_row = 1093
        if category == 'ERROR_MT':
            source_start_row = 1193
            source_end_row = 1195
        if category == 'ERROR_OUT':
            source_start_row = 1195
            source_end_row = 1213

        col_name = self.df.columns[source_col]

        # ✅ 데이터 자르기
        df_partial = self.df.loc[:, [col_name]].iloc[source_start_row:source_end_row]

        # 헤더 출력
        ws.cell(row=target_start_row, column=target_col, value=col_name)

        # ✅ 실제 데이터 출력 (예외 방지)
        for i, (_, row) in enumerate(df_partial.iterrows(), start=target_start_row):
            value = row[col_name]

            # Series나 리스트일 경우 첫 값만 사용
            if isinstance(value, (pd.Series, list)):
                value = value[0] if len(value) > 0 else ""

            # NaN 값은 공백으로 처리
            if pd.isna(value):
                value = ""

            # Excel에 출력
            ws.cell(row=i, column=target_col, value=value)

        last_row = target_start_row + len(df_partial)
        print(f"✅ 복사된 범위: 시트 기준 {target_start_row}행 ~ {last_row}행, 열 {target_col}")
        return last_row



    
def add_to_address(addr_str, offset):
    
    match = re.match(r'([A-Za-z]+)(\d+)', addr_str)
    if not match:
        raise ValueError(f"형식이 올바르지 않습니다 (예: 'M1700' {addr_str})")

    prefix, number = match.groups()
    new_number = int(number) + offset
    return f"{prefix}{new_number}"
def add_to_bit_address(addr_str, bit_offset):

    # 예: D600.0 → prefix=D, word=600, bit=0
    match = re.match(r'([A-Za-z]+)(\d+)\.(\d+)', addr_str)
    if not match:
        raise ValueError("형식이 잘못되었습니다 (예: 'D600.0')")

    prefix, word_str, bit_str = match.groups()
    word = int(word_str)
    bit = int(bit_str)

    # 전체 비트 주소로 환산
    total_bit = word * 16 + bit + bit_offset

    # 다시 word와 bit 분리
    new_word = total_bit // 16
    new_bit = total_bit % 16

    # bit를 16진수로 표현 (A, B, ..., F)
    hex_bit = format(new_bit, 'X')

    return f"{prefix}{new_word}.{hex_bit}"
def add_bit_offset(address_str: str, bit_offset: int) -> str:
    """
    예: D700, 1 → D700.1
        D700, 10 → D700.A
    """
    match = re.match(r'([A-Za-z]+)(\d+)', address_str)
    if not match:
        raise ValueError("형식이 올바르지 않습니다 (예: 'D700')")

    prefix, number_str = match.groups()
    word = int(number_str)

    # 16진수로 비트 오프셋 계산
    bit_hex = format(bit_offset, 'X')  # 10 → 'A', 15 → 'F' 등

    return f"{prefix}{word}.{bit_hex}"
def add_hex_offset(addr_str, offset):
    addr_str = addr_str.strip()

    # 1. 비트 주소가 포함된 경우는 제외
    if '.' in addr_str:
        raise ValueError(f"비트 주소는 처리하지 않습니다: {addr_str}")

    # 2. X/Y는 16진수 처리
    match = re.match(r"^([XY])([0-9A-Fa-f]+)$", addr_str)
    if match:
        prefix, hex_str = match.groups()
        base = int(hex_str, 16)
        result = base + offset
        return f"{prefix}{format(result, 'X').upper()}"

    # 3. 나머지(ZR, D 등)는 10진수 처리
    match = re.match(r"^([A-Za-z]+)([0-9]+)$", addr_str)
    if match:
        prefix, dec_str = match.groups()
        base = int(dec_str)
        result = base + offset
        return f"{prefix}{result}"

    # 4. 포맷이 맞지 않으면 예외
    raise ValueError(f"입력 형식이 잘못되었습니다. 받은값: {addr_str} (예: 'X180', 'ZR60000')")
#엑셀 파일 경로
#excel_file = 'PCB_UNIT_MECRO_MAP_20250608_1.xlsx'
#unit_start_adr = UnitStartAdrLookup(excel_file)



#get_m_start_adr= unit_start_adr.get_start_adr(3,"OUTPUT","MT_START")



"""
# 엑셀 파일 경로
excel_file = 'PCB_MOTION_MECRO_MAP_20250608_1.xlsx'

# 클래스 인스턴스 생성
motor_prog = MotorProgLookup(excel_file)
motor_pos_adr = MotorPosDataLookup(excel_file)
motor_param_adr = MotorParamDataLookup(excel_file)
motor_start_adr = MotorStartAdrLookup(excel_file)
motor_motion_recv_adr = MotionAdrRecvLookup(excel_file)
motor_motion_send_adr = MotionAdrSendLookup(excel_file)
ladder_mecro = LadderMecro(excel_file)

# 전체 데이터 출력
motor_prog.print_data()
program_no= 17
program_count= 1
pos_no = 1
# 모션 CPU Axis status RECV 주소에서 사용하는 비트
axis_status_recv_bit_no = 0
axis_monitor_recv_word_no = 0
# 모션 CPU Axis status Send 주소에서 사용하는 비트
axis_command_send_bit_no = 0
axis_position_send_word_no = 0
"""



# 특정 값 조회
"""
axis_no = motor_prog.get_axis_no(program_no, program_count)
axis_fix_pos_adr = motor_pos_adr.get_pos_adr('FIXED',axis_no)
axis_offs_pos_adr = motor_pos_adr.get_pos_adr('OFFSET',axis_no)
axis_model_pos_adr = motor_pos_adr.get_pos_adr('MODEL',axis_no)
axis_machine_pos_adr = motor_pos_adr.get_pos_adr('MACHINE',axis_no)
axis_vision_pos_adr = motor_pos_adr.get_pos_adr('VISION',axis_no)
axis_target_pos_adr = motor_pos_adr.get_pos_adr('TARGET',axis_no)
axis_param_highspeed_adr = motor_param_adr.get_param_adr('HIGHSPEED',axis_no)
axis_param_lowspeed_adr = motor_param_adr.get_param_adr('LOWSPEED',axis_no)
axis_param_acctime_adr = motor_param_adr.get_param_adr('ACCTIME',axis_no)
axis_param_dectime_adr = motor_param_adr.get_param_adr('DECTIME',axis_no)
axis_param_joglowspd_adr = motor_param_adr.get_param_adr('JOGLOWSPD',axis_no)
axis_param_joghighspd_adr = motor_param_adr.get_param_adr('JOGHIGHSPD',axis_no)
axis_param_softlimitp_adr = motor_param_adr.get_param_adr('SOFTLIMIT+',axis_no)
axis_param_softlimitm_adr = motor_param_adr.get_param_adr('SOFTLIMIT-',axis_no)
axis_m_start_adr = motor_start_adr.get_start_adr('M_START',program_no)
axis_d_start_adr = motor_start_adr.get_start_adr('D_START',program_no)
axis_t_start_adr = motor_start_adr.get_start_adr('T_START',program_no)
# 모션 CPU RECV 주소 조회
motion_axis_status_recv_adr = motor_motion_recv_adr.get_motion_axis_status_recv_adr('START',axis_no,axis_status_recv_bit_no)
motion_axis_monitor_recv_adr = motor_motion_recv_adr.get_motion_axis_monitor_recv_adr('START',axis_no,axis_monitor_recv_word_no)
motion_axis_SFC_home_recv_adr = motor_motion_recv_adr.get_motion_axis_SFC_home_recv_adr('START',axis_no,axis_no)
motion_axis_SFC_pos_recv_adr = motor_motion_recv_adr.get_motion_axis_SFC_pos_recv_adr('START',axis_no,axis_no)
motion_axis_SFC_axis_recv_adr = motor_motion_recv_adr.get_motion_axis_SFC_axis_recv_adr('START',axis_no,axis_no)
# 모션 CPU SEND 주소 조회
motion_axis_jogspeed_send_adr = motor_motion_send_adr.get_motion_axis_jogspeed_send_adr('START',axis_no,axis_no)
motion_axis_Command_send_adr = motor_motion_send_adr.get_motion_axis_Command_send_adr('START',axis_no,axis_command_send_bit_no)
motion_axis_position_send_adr = motor_motion_send_adr.get_motion_axis_position_send_adr('START',axis_no,axis_position_send_word_no)
motion_axis_SFC_home_Send_adr = motor_motion_send_adr.get_motion_axis_SFC_home_Send_adr('START',axis_no,axis_no)
motion_axis_SFC_pos_Send_adr = motor_motion_send_adr.get_motion_axis_SFC_pos_Send_adr('START',axis_no,axis_no)



print(f'axis_no:{axis_no}')
print(f'axis_fix_pos_adr:{axis_fix_pos_adr}')
print(f'axis_offs_pos_adr:{axis_offs_pos_adr}')
print(f'axis_model_pos_adr:{axis_model_pos_adr}')
print(f'axis_machine_pos_adr:{axis_machine_pos_adr}')
print(f'axis_vision_pos_adr:{axis_vision_pos_adr}')
print(f'axis_target_pos_adr:{axis_target_pos_adr}')
print(f'axis_param_highspeed_adr:{axis_param_highspeed_adr}')
print(f'axis_param_lowspeed_adr:{axis_param_lowspeed_adr}')
print(f'axis_param_acctime_adr:{axis_param_acctime_adr}')
print(f'axis_param_joglowspeed_adr:{axis_param_joglowspd_adr}')
print(f'axis_param_softlimit+_adr:{axis_param_softlimitp_adr}')
print(f'axis_param_softlimit-_adr:{axis_param_softlimitm_adr}')
print(f'axis_m_start_adr:{axis_m_start_adr}')
print(f'axis_d_start_adr:{axis_d_start_adr}')
print(f'axis_t_start_adr:{axis_t_start_adr}')
print(f'motion_axis_status_recv_adr:{motion_axis_status_recv_adr}')
print(f'motion_axis_monitor_recv_adr:{motion_axis_monitor_recv_adr}')
print(f'motion_axis_SFC_home_recv_adr:{motion_axis_SFC_home_recv_adr}')
print(f'motion_axis_SFC_pos_recv_adr:{motion_axis_SFC_pos_recv_adr}')
print(f'motion_axis_SFC_axis_recv_adr:{motion_axis_SFC_axis_recv_adr}')
print(f'motion_axis_jogspeed_send_adr:{motion_axis_jogspeed_send_adr}')
print(f'motion_axis_Command_send_adr:{motion_axis_Command_send_adr}')
print(f'motion_axis_position_send_adr:{motion_axis_position_send_adr}')
print(f'motion_axis_SFC_home_Send_adr:{motion_axis_SFC_home_Send_adr}')
print(f'motion_axis_SFC_pos_Send_adr:{motion_axis_SFC_pos_Send_adr}')
axis_m_start_adr = add_to_address(axis_m_start_adr, 1)
print(f'axis_m_start_adr+ 1:{axis_m_start_adr}')
print(f'motion_axis_SFC_home_recv_adr:{motion_axis_SFC_home_recv_adr}')
motion_axis_SFC_home_recv_adr = add_to_bit_address(motion_axis_SFC_home_recv_adr,1)
print(f'motion_axis_SFC_home_recv_adr + 1:{motion_axis_SFC_home_recv_adr}')
"""
